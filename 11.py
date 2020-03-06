import os
from openpyxl import Workbook
from openpyxl import load_workbook

wb2 = load_workbook(r'C:\Users\gaopeihan\Desktop\ppp.xlsx')
wc2 = wb2.worksheets[0]

def excel_(path,row1):
    wb = load_workbook(r'D:\qq接收文件\电气1601健康卡和体温统计表\电气1601健康卡\\' + str(path))
    wc = wb.worksheets[0]
    i = 2
    for x in range(3,33,2):
        c1 = wc2.cell(row=row1,column=x).value
        c2 = wc2.cell(row=row1,column=x+1).value
        wc.cell(row=14,column=i,value=c2)
        wc.cell(row=13,column=i,value=c1)
        i = i+1
    i = 2
    for x in range(33,43,2):
        c1 = wc2.cell(row=row1,column=x).value
        c2 = wc2.cell(row=row1,column=x+1).value
        wc.cell(row=17,column=i,value=c2)
        wc.cell(row=16,column=i,value=c1)
        i = i+1   
    wc.cell(row=18,column=3,value='许鸥')
    wc.cell(row=18,column=7,value='辅导员')
    wc.cell(row=18,column=13,value='13513360880')
    wb.save(r'D:\qq接收文件\电气1601健康卡和体温统计表\电气1601健康卡\\' + path)
    print(path)


def list_dir(path):
    list_ = os.listdir(path)
    for list1 in list_:
        row = int(list1[0:2]) + 3
        excel_(list1,row)
        
             
if __name__ == "__main__":
    list_dir(r'D:\qq接收文件\电气1601健康卡和体温统计表\电气1601健康卡')